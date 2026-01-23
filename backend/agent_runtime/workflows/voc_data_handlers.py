"""
VoC 데이터 핸들러

다양한 형식의 VoC 데이터를 파싱하고 정규화하는 핸들러
- CSV
- Excel (openpyxl)
- API 데이터
- 텍스트
"""

import csv
import io
from abc import ABC, abstractmethod
from dataclasses import dataclass
from typing import Any

import structlog

logger = structlog.get_logger()


@dataclass
class VoCRecord:
    """정규화된 VoC 레코드"""

    text: str  # VoC 내용
    category: str | None = None  # 카테고리 (있는 경우)
    timestamp: str | None = None  # 시간 (있는 경우)
    customer_id: str | None = None  # 고객 ID (있는 경우)
    severity: str | None = None  # 심각도 (있는 경우)
    metadata: dict[str, Any] | None = None  # 추가 메타데이터


class BaseVoCHandler(ABC):
    """VoC 데이터 핸들러 기본 클래스"""

    def __init__(self):
        self.logger = logger.bind(handler=self.__class__.__name__)

    @abstractmethod
    def parse(self, data: Any) -> list[VoCRecord]:
        """데이터 파싱 및 정규화"""
        pass

    def validate(self, records: list[VoCRecord]) -> list[VoCRecord]:
        """레코드 유효성 검증"""
        valid_records = []
        for record in records:
            if record.text and len(record.text.strip()) > 0:
                valid_records.append(record)
        return valid_records


class CSVVoCHandler(BaseVoCHandler):
    """CSV 파일 VoC 핸들러

    지원 컬럼:
    - text/content/description/voc: VoC 내용 (필수)
    - category/type/분류: 카테고리
    - timestamp/date/created_at/날짜: 시간
    - customer_id/customer/고객: 고객 ID
    - severity/priority/심각도: 심각도
    """

    TEXT_COLUMNS = ["text", "content", "description", "voc", "내용", "본문"]
    CATEGORY_COLUMNS = ["category", "type", "분류", "카테고리"]
    TIMESTAMP_COLUMNS = ["timestamp", "date", "created_at", "날짜", "일시"]
    CUSTOMER_COLUMNS = ["customer_id", "customer", "고객", "고객id"]
    SEVERITY_COLUMNS = ["severity", "priority", "심각도", "중요도"]

    def parse(self, data: bytes | str) -> list[VoCRecord]:
        """CSV 데이터 파싱"""
        records = []

        # bytes → str 변환
        if isinstance(data, bytes):
            # UTF-8 시도, 실패 시 CP949 (한글 Windows)
            try:
                content = data.decode("utf-8")
            except UnicodeDecodeError:
                content = data.decode("cp949")
        else:
            content = data

        # CSV 파싱
        reader = csv.DictReader(io.StringIO(content))

        # 컬럼 매핑 찾기
        fieldnames = [f.lower() for f in (reader.fieldnames or [])]
        text_col = self._find_column(fieldnames, self.TEXT_COLUMNS)
        category_col = self._find_column(fieldnames, self.CATEGORY_COLUMNS)
        timestamp_col = self._find_column(fieldnames, self.TIMESTAMP_COLUMNS)
        customer_col = self._find_column(fieldnames, self.CUSTOMER_COLUMNS)
        severity_col = self._find_column(fieldnames, self.SEVERITY_COLUMNS)

        if not text_col:
            self.logger.warning("No text column found in CSV", fieldnames=fieldnames)
            return []

        for row in reader:
            # 대소문자 무시하고 값 가져오기
            row_lower = {k.lower(): v for k, v in row.items()}

            text = row_lower.get(text_col, "").strip()
            if not text:
                continue

            record = VoCRecord(
                text=text,
                category=row_lower.get(category_col, "").strip() if category_col else None,
                timestamp=row_lower.get(timestamp_col, "").strip() if timestamp_col else None,
                customer_id=row_lower.get(customer_col, "").strip() if customer_col else None,
                severity=row_lower.get(severity_col, "").strip() if severity_col else None,
                metadata={k: v for k, v in row_lower.items() if k not in [text_col]},
            )
            records.append(record)

        self.logger.info("CSV parsed", total_records=len(records))
        return self.validate(records)

    def _find_column(self, fieldnames: list[str], candidates: list[str]) -> str | None:
        """컬럼 이름 찾기"""
        for candidate in candidates:
            if candidate.lower() in fieldnames:
                return candidate.lower()
        return None


class ExcelVoCHandler(BaseVoCHandler):
    """Excel 파일 VoC 핸들러

    openpyxl 사용
    """

    TEXT_COLUMNS = ["text", "content", "description", "voc", "내용", "본문"]
    CATEGORY_COLUMNS = ["category", "type", "분류", "카테고리"]
    TIMESTAMP_COLUMNS = ["timestamp", "date", "created_at", "날짜", "일시"]
    CUSTOMER_COLUMNS = ["customer_id", "customer", "고객", "고객id"]
    SEVERITY_COLUMNS = ["severity", "priority", "심각도", "중요도"]

    def parse(self, data: bytes) -> list[VoCRecord]:
        """Excel 데이터 파싱"""
        try:
            from openpyxl import load_workbook
        except ImportError as err:
            self.logger.error("openpyxl not installed")
            raise ImportError("openpyxl이 설치되어 있지 않습니다. pip install openpyxl") from err

        records = []

        # 바이트 데이터에서 워크북 로드
        wb = load_workbook(filename=io.BytesIO(data), read_only=True)
        ws = wb.active

        if ws is None:
            self.logger.warning("No active worksheet")
            return []

        # 헤더 행 읽기
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).lower() if h else "" for h in rows[0]]

        # 컬럼 인덱스 찾기
        text_idx = self._find_column_index(headers, self.TEXT_COLUMNS)
        category_idx = self._find_column_index(headers, self.CATEGORY_COLUMNS)
        timestamp_idx = self._find_column_index(headers, self.TIMESTAMP_COLUMNS)
        customer_idx = self._find_column_index(headers, self.CUSTOMER_COLUMNS)
        severity_idx = self._find_column_index(headers, self.SEVERITY_COLUMNS)

        if text_idx is None:
            self.logger.warning("No text column found in Excel", headers=headers)
            return []

        # 데이터 행 처리
        for row in rows[1:]:
            text = str(row[text_idx]).strip() if row[text_idx] else ""
            if not text:
                continue

            record = VoCRecord(
                text=text,
                category=(
                    str(row[category_idx]).strip()
                    if category_idx is not None and row[category_idx]
                    else None
                ),
                timestamp=(
                    str(row[timestamp_idx]).strip()
                    if timestamp_idx is not None and row[timestamp_idx]
                    else None
                ),
                customer_id=(
                    str(row[customer_idx]).strip()
                    if customer_idx is not None and row[customer_idx]
                    else None
                ),
                severity=(
                    str(row[severity_idx]).strip()
                    if severity_idx is not None and row[severity_idx]
                    else None
                ),
            )
            records.append(record)

        wb.close()
        self.logger.info("Excel parsed", total_records=len(records))
        return self.validate(records)

    def _find_column_index(self, headers: list[str], candidates: list[str]) -> int | None:
        """컬럼 인덱스 찾기"""
        for candidate in candidates:
            if candidate.lower() in headers:
                return headers.index(candidate.lower())
        return None


class APIVoCHandler(BaseVoCHandler):
    """API 데이터 VoC 핸들러

    JSON 형식의 API 응답 데이터 처리
    """

    def parse(self, data: list[dict[str, Any]]) -> list[VoCRecord]:
        """API 데이터 파싱"""
        records = []

        for item in data:
            # 다양한 키 이름 지원
            text = (
                item.get("text")
                or item.get("content")
                or item.get("description")
                or item.get("message")
                or item.get("voc")
                or ""
            )

            if not text:
                continue

            record = VoCRecord(
                text=str(text).strip(),
                category=item.get("category") or item.get("type"),
                timestamp=item.get("timestamp") or item.get("created_at") or item.get("date"),
                customer_id=item.get("customer_id") or item.get("customer"),
                severity=item.get("severity") or item.get("priority"),
                metadata={k: v for k, v in item.items() if k not in ["text", "content"]},
            )
            records.append(record)

        self.logger.info("API data parsed", total_records=len(records))
        return self.validate(records)


class TextVoCHandler(BaseVoCHandler):
    """텍스트 VoC 핸들러

    줄바꿈으로 구분된 텍스트 처리
    """

    def parse(self, data: str) -> list[VoCRecord]:
        """텍스트 데이터 파싱"""
        records = []

        # 줄바꿈으로 분리
        lines = data.strip().split("\n")

        for line in lines:
            text = line.strip()
            if not text:
                continue

            # 빈 줄이나 너무 짧은 줄 제외
            if len(text) < 5:
                continue

            record = VoCRecord(text=text)
            records.append(record)

        self.logger.info("Text parsed", total_records=len(records))
        return self.validate(records)


def get_handler(source_type: str) -> BaseVoCHandler:
    """소스 타입에 맞는 핸들러 반환

    Args:
        source_type: "csv", "excel", "api", "text"

    Returns:
        BaseVoCHandler 인스턴스

    Raises:
        ValueError: 지원하지 않는 소스 타입
    """
    handlers = {
        "csv": CSVVoCHandler,
        "excel": ExcelVoCHandler,
        "api": APIVoCHandler,
        "text": TextVoCHandler,
    }

    handler_class = handlers.get(source_type.lower())
    if handler_class is None:
        raise ValueError(f"지원하지 않는 소스 타입: {source_type}. 지원: {list(handlers.keys())}")

    # NOTE: handlers의 값은 구체적인 서브클래스이므로 인스턴스화 가능
    return handler_class()  # type: ignore[abstract]
