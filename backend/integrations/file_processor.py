"""
파일 처리 유틸리티

이미지, PDF, 문서, 텍스트 파일에서 세미나 정보 추출
"""

import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import httpx
import structlog

logger = structlog.get_logger()


@dataclass
class SeminarInfo:
    """세미나 정보"""

    title: str
    description: str | None = None
    date: str | None = None
    organizer: str | None = None
    url: str | None = None
    categories: list[str] = field(default_factory=list)
    confidence: float = 0.8

    def to_dict(self) -> dict:
        """딕셔너리 변환"""
        return {
            "title": self.title,
            "description": self.description,
            "date": self.date,
            "organizer": self.organizer,
            "url": self.url,
            "categories": self.categories,
            "confidence": self.confidence,
        }


class FileProcessor:
    """
    파일에서 세미나 정보 추출

    지원 형식:
    - 이미지: jpg, png, webp (Claude Vision API)
    - PDF: pdf (pypdf + LLM)
    - 문서: docx, xlsx (python-docx, openpyxl)
    - 텍스트: txt, csv, json, md (직접 파싱)
    """

    # 지원 파일 확장자
    IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    PDF_EXTENSIONS = {".pdf"}
    DOCUMENT_EXTENSIONS = {".docx", ".xlsx", ".xls"}
    TEXT_EXTENSIONS = {".txt", ".csv", ".json", ".md", ".markdown"}

    def __init__(self):
        self.anthropic_api_key = self._get_anthropic_api_key()

    def _get_anthropic_api_key(self) -> str | None:
        """Anthropic API 키 가져오기"""
        import os

        return os.getenv("ANTHROPIC_API_KEY")

    def _get_file_extension(self, filename: str) -> str:
        """파일 확장자 추출"""
        return Path(filename).suffix.lower()

    async def process_file(
        self, content: bytes, filename: str, content_type: str
    ) -> list[SeminarInfo]:
        """
        파일에서 세미나 정보 추출

        Args:
            content: 파일 내용
            filename: 파일명
            content_type: MIME 타입

        Returns:
            추출된 세미나 정보 목록
        """
        ext = self._get_file_extension(filename)

        if ext in self.IMAGE_EXTENSIONS:
            return await self.process_image(content, content_type)
        elif ext in self.PDF_EXTENSIONS:
            return await self.process_pdf(content)
        elif ext in self.DOCUMENT_EXTENSIONS:
            return await self.process_document(content, ext)
        elif ext in self.TEXT_EXTENSIONS:
            return await self.process_text_file(content, ext)
        else:
            logger.warning(f"지원하지 않는 파일 형식: {ext}")
            return []

    async def process_image(self, content: bytes, content_type: str) -> list[SeminarInfo]:
        """
        이미지 OCR → 세미나 정보

        Claude Vision API를 사용하여 이미지에서 텍스트 추출
        """
        if not self.anthropic_api_key:
            logger.warning("ANTHROPIC_API_KEY가 설정되지 않았습니다")
            return []

        import base64

        # base64 인코딩
        image_data = base64.standard_b64encode(content).decode("utf-8")

        # MIME 타입 결정
        media_type = content_type
        if not media_type or media_type == "application/octet-stream":
            media_type = "image/jpeg"

        # Claude Vision API 호출
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key": self.anthropic_api_key,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json",
                    },
                    json={
                        "model": "claude-sonnet-4-20250514",
                        "max_tokens": 2000,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {
                                        "type": "image",
                                        "source": {
                                            "type": "base64",
                                            "media_type": media_type,
                                            "data": image_data,
                                        },
                                    },
                                    {
                                        "type": "text",
                                        "text": self._get_extraction_prompt(),
                                    },
                                ],
                            }
                        ],
                    },
                    timeout=60.0,
                )

                if response.status_code != 200:
                    logger.error(
                        "Claude Vision API 오류",
                        status=response.status_code,
                        body=response.text,
                    )
                    return []

                result = response.json()
                text_content = result["content"][0]["text"]
                return self._parse_llm_response(text_content)

        except Exception as e:
            logger.error("이미지 처리 오류", error=str(e))
            return []

    async def process_pdf(self, content: bytes) -> list[SeminarInfo]:
        """
        PDF → 세미나 정보

        pypdf로 텍스트 추출 후 LLM으로 분석
        """
        try:
            from io import BytesIO

            # pypdf 동적 임포트
            try:
                from pypdf import PdfReader
            except ImportError:
                logger.warning("pypdf가 설치되지 않았습니다. pip install pypdf 실행 필요")
                return []

            reader = PdfReader(BytesIO(content))
            text = ""
            for page in reader.pages[:5]:  # 최대 5페이지만 처리
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

            if not text.strip():
                logger.warning("PDF에서 텍스트를 추출할 수 없습니다")
                return []

            # LLM으로 세미나 정보 추출
            return await self._extract_with_llm(text)

        except Exception as e:
            logger.error("PDF 처리 오류", error=str(e))
            return []

    async def process_document(self, content: bytes, extension: str) -> list[SeminarInfo]:
        """
        DOCX/XLSX → 세미나 정보
        """
        try:
            from io import BytesIO

            if extension == ".docx":
                # python-docx 동적 임포트
                try:
                    from docx import Document
                except ImportError:
                    logger.warning(
                        "python-docx가 설치되지 않았습니다. pip install python-docx 실행 필요"
                    )
                    return []

                doc = Document(BytesIO(content))
                text = "\n".join([para.text for para in doc.paragraphs])

            elif extension in (".xlsx", ".xls"):
                # openpyxl 동적 임포트
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    logger.warning("openpyxl이 설치되지 않았습니다. pip install openpyxl 실행 필요")
                    return []

                wb = load_workbook(BytesIO(content))
                ws = wb.active
                if ws is None:
                    return []

                rows = []
                for row in ws.iter_rows(max_row=100):  # 최대 100행
                    row_data = [str(cell.value) if cell.value else "" for cell in row]
                    rows.append(",".join(row_data))
                text = "\n".join(rows)
            else:
                return []

            if not text.strip():
                return []

            return await self._extract_with_llm(text)

        except Exception as e:
            logger.error("문서 처리 오류", error=str(e))
            return []

    async def process_text_file(self, content: bytes, extension: str) -> list[SeminarInfo]:
        """
        텍스트 파일 → 세미나 정보
        """
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            try:
                text = content.decode("cp949")  # 한글 Windows 인코딩
            except UnicodeDecodeError:
                logger.warning("텍스트 디코딩 실패")
                return []

        if extension == ".json":
            return self._parse_json(text)
        elif extension == ".csv":
            return self._parse_csv(text)
        else:
            return await self._extract_with_llm(text)

    async def process_text(self, text: str) -> list[SeminarInfo]:
        """
        텍스트에서 세미나 정보 추출

        URL 감지 및 웹 페이지 스크래핑
        """
        if not text.strip():
            return []

        # URL 추출
        urls = re.findall(r"https?://[^\s<>\"']+", text)

        results: list[SeminarInfo] = []

        # URL에서 세미나 정보 추출
        for url in urls[:5]:  # 최대 5개 URL만 처리
            seminar = await self._extract_from_url(url)
            if seminar:
                results.append(seminar)

        # URL이 없거나 추출 실패 시 텍스트 자체 분석
        if not results and len(text) > 20:
            results = await self._extract_with_llm(text)

        return results

    async def _extract_from_url(self, url: str) -> SeminarInfo | None:
        """
        URL에서 세미나 정보 추출
        """
        try:
            # 알려진 세미나 플랫폼 체크
            if "festa.io" in url or "event-us.kr" in url or "onoffmix.com" in url:
                return await self._scrape_seminar_page(url)

            # 일반 URL은 OG 태그 추출
            return await self._extract_og_tags(url)

        except Exception as e:
            logger.warning(f"URL 처리 실패: {url}", error=str(e))
            return None

    async def _scrape_seminar_page(self, url: str) -> SeminarInfo | None:
        """
        세미나 플랫폼 페이지 스크래핑
        """
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(
                    url,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                    },
                    follow_redirects=True,
                    timeout=10.0,
                )

                if response.status_code != 200:
                    return None

                html = response.text

                # 기본 OG 태그 추출
                title = self._extract_meta(html, "og:title") or self._extract_title(html)
                description = self._extract_meta(html, "og:description")

                if not title:
                    return None

                return SeminarInfo(
                    title=title,
                    description=description,
                    url=url,
                    confidence=0.7,
                )

        except Exception as e:
            logger.warning(f"페이지 스크래핑 실패: {url}", error=str(e))
            return None

    async def _extract_og_tags(self, url: str) -> SeminarInfo | None:
        """
        Open Graph 태그 추출
        """
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(
                    url,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                    },
                    follow_redirects=True,
                    timeout=10.0,
                )

                if response.status_code != 200:
                    return None

                html = response.text
                title = self._extract_meta(html, "og:title") or self._extract_title(html)
                description = self._extract_meta(html, "og:description")

                if not title:
                    return None

                return SeminarInfo(
                    title=title,
                    description=description,
                    url=url,
                    confidence=0.6,
                )

        except Exception as e:
            logger.warning(f"OG 태그 추출 실패: {url}", error=str(e))
            return None

    def _extract_meta(self, html: str, property_name: str) -> str | None:
        """
        메타 태그 추출
        """
        patterns = [
            rf'<meta[^>]+property=["\']?{property_name}["\']?[^>]+content=["\']([^"\']+)["\']',
            rf'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']?{property_name}["\']?',
        ]

        for pattern in patterns:
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                return match.group(1).strip()

        return None

    def _extract_title(self, html: str) -> str | None:
        """
        <title> 태그 추출
        """
        match = re.search(r"<title>([^<]+)</title>", html, re.IGNORECASE)
        return match.group(1).strip() if match else None

    async def _extract_with_llm(self, text: str) -> list[SeminarInfo]:
        """
        LLM을 사용하여 텍스트에서 세미나 정보 추출
        """
        if not self.anthropic_api_key:
            logger.warning("ANTHROPIC_API_KEY가 설정되지 않았습니다")
            return self._extract_basic(text)

        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key": self.anthropic_api_key,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json",
                    },
                    json={
                        "model": "claude-sonnet-4-20250514",
                        "max_tokens": 2000,
                        "messages": [
                            {
                                "role": "user",
                                "content": f"{self._get_extraction_prompt()}\n\n텍스트:\n{text[:3000]}",
                            }
                        ],
                    },
                    timeout=30.0,
                )

                if response.status_code != 200:
                    logger.error(
                        "Claude API 오류",
                        status=response.status_code,
                        body=response.text,
                    )
                    return self._extract_basic(text)

                result = response.json()
                text_content = result["content"][0]["text"]
                return self._parse_llm_response(text_content)

        except Exception as e:
            logger.error("LLM 추출 오류", error=str(e))
            return self._extract_basic(text)

    def _get_extraction_prompt(self) -> str:
        """
        세미나 정보 추출 프롬프트
        """
        return """이 이미지 또는 텍스트에서 세미나/이벤트/컨퍼런스 정보를 추출해주세요.

다음 JSON 형식으로 응답해주세요:
```json
{
  "seminars": [
    {
      "title": "세미나 제목",
      "description": "세미나 설명 (선택)",
      "date": "YYYY-MM-DD 형식의 날짜 (선택)",
      "organizer": "주최자/주관사 (선택)",
      "url": "관련 URL (선택)",
      "categories": ["AI", "기술"] // 관련 카테고리
    }
  ]
}
```

세미나 정보가 없으면 {"seminars": []}를 반환해주세요.
JSON만 반환하고 다른 설명은 생략해주세요."""

    def _parse_llm_response(self, response: str) -> list[SeminarInfo]:
        """
        LLM 응답 파싱
        """
        try:
            # JSON 블록 추출
            json_match = re.search(r"```json\s*(.*?)\s*```", response, re.DOTALL)
            json_str = json_match.group(1) if json_match else response

            data = json.loads(json_str)
            seminars_data = data.get("seminars", [])

            results = []
            for s in seminars_data:
                if isinstance(s, dict) and s.get("title"):
                    results.append(
                        SeminarInfo(
                            title=s["title"],
                            description=s.get("description"),
                            date=s.get("date"),
                            organizer=s.get("organizer"),
                            url=s.get("url"),
                            categories=s.get("categories", []),
                            confidence=0.85,
                        )
                    )

            return results

        except (json.JSONDecodeError, KeyError) as e:
            logger.warning("LLM 응답 파싱 실패", error=str(e), response=response[:200])
            return []

    def _extract_basic(self, text: str) -> list[SeminarInfo]:
        """
        기본 텍스트 추출 (LLM 없이)
        """
        # 제목 추정 (첫 줄 또는 길이가 적당한 줄)
        lines = [line.strip() for line in text.split("\n") if line.strip()]
        if not lines:
            return []

        title = lines[0][:100] if lines else "제목 없음"

        # 날짜 패턴 추출
        date_patterns = [
            r"(\d{4}[-./]\d{1,2}[-./]\d{1,2})",
            r"(\d{1,2}월\s*\d{1,2}일)",
        ]
        date = None
        for pattern in date_patterns:
            match = re.search(pattern, text)
            if match:
                date = match.group(1)
                break

        return [
            SeminarInfo(
                title=title,
                description=text[:200] if len(text) > 50 else None,
                date=date,
                confidence=0.5,
            )
        ]

    def _parse_json(self, text: str) -> list[SeminarInfo]:
        """
        JSON 파싱
        """
        try:
            data = json.loads(text)

            # 배열인 경우
            if isinstance(data, list):
                items = data
            # 객체인 경우
            elif isinstance(data, dict):
                items = data.get("seminars", data.get("events", [data]))
            else:
                return []

            results = []
            for item in items:
                if isinstance(item, dict):
                    title = item.get("title") or item.get("name") or item.get("제목")
                    if title:
                        results.append(
                            SeminarInfo(
                                title=title,
                                description=item.get("description") or item.get("설명"),
                                date=item.get("date") or item.get("날짜"),
                                organizer=item.get("organizer")
                                or item.get("주최")
                                or item.get("host"),
                                url=item.get("url") or item.get("link"),
                                categories=item.get("categories", []),
                                confidence=0.9,
                            )
                        )

            return results

        except json.JSONDecodeError:
            return []

    def _parse_csv(self, text: str) -> list[SeminarInfo]:
        """
        CSV 파싱
        """
        import csv
        from io import StringIO

        try:
            reader = csv.DictReader(StringIO(text))
            results = []

            for row in reader:
                # 제목 필드 찾기
                title = None
                for key in ["title", "name", "제목", "세미나명", "이벤트명"]:
                    if key in row and row[key]:
                        title = row[key]
                        break

                if title:
                    results.append(
                        SeminarInfo(
                            title=title,
                            description=row.get("description") or row.get("설명"),
                            date=row.get("date") or row.get("날짜"),
                            organizer=row.get("organizer") or row.get("주최") or row.get("host"),
                            url=row.get("url") or row.get("link"),
                            categories=[],
                            confidence=0.9,
                        )
                    )

            return results

        except Exception:
            return []
